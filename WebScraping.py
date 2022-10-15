import json
import certifi
import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from lxml import html
from openpyxl import load_workbook


def read_spreadsheet(file, sheet_name):
    pd.options.display.float_format = '{:.0f}'.format
    wb = load_workbook(filename=file)

    utils = wb[sheet_name]
    stocks = []

    for row in utils.iter_rows(min_row=2):
        stock = []
        for cell in row:
            stock.append(cell.value)
        stocks.append(stock)

    wb.close()
    return stocks


def get_request(url):
    http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED', ca_certs=certifi.where())
    return http.request('GET', url, preload_content=False)


def get_summary(ticker):
    url = "https://finance.yahoo.com/quote/%s?p=%s" % (ticker, ticker)
    request = get_request(url)
    parser = html.fromstring(request.data)
    summary_table = parser.xpath('//div[contains(@data-test,"summary-table")]//tr')
    soup = BeautifulSoup(request, 'lxml')

    for x in summary_table:
        print(x.xpath('.//td[1]//text()'), x.xpath('.//td[2]//text()'))


def get_stats(ticker):
    url = "https://finance.yahoo.com/quote/%s/key-statistics?p=%s" % (ticker, ticker)
    request = get_request(url)
    parser = html.fromstring(request.data)
    soup = BeautifulSoup(request, 'lxml')
    stat_table = parser.xpath('//div[contains(@data-test, "qsp-statistics")]//tr')

    for stat in stat_table:
        print(stat.xpath('.//td[1]//text()'), stat.xpath('.//td[2]//text()'))


stock_list = read_spreadsheet('Spreadsheets/Book1.xlsx', 'Utilities')

get_summary(stock_list[0][1])
