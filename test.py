import certifi
import pandas as pd
import urllib3
from bs4 import BeautifulSoup
from lxml import html
from openpyxl import load_workbook, Workbook


def get_request(url):
    http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED', ca_certs=certifi.where())
    return http.request('GET', url, preload_content=False)


def get_summary(stock_info):
    url = "https://finance.yahoo.com/quote/%s?p=%s" % (stock_info[1], stock_info[1])
    request = get_request(url)
    parser = html.fromstring(request.data)
    summary_table = parser.xpath('//div[contains(@data-test,"summary-table")]//tr')

    for x in summary_table:
        stock_info.append(x.xpath('.//td[2]//text()')[0])


def get_stats(stock_info):
    url = "https://finance.yahoo.com/quote/%s/key-statistics?p=%s" % (stock_info, stock_info)
    request = get_request(url)
    parser = html.fromstring(request.data)
    stat_table = parser.xpath('//div[contains(@data-test, "qsp-statistics")]//tr')

    for stat in stat_table:
        stock_info.append(stat.xpath('.//td[2]//text()'))


def web_scrape(spreadsheet, industry, cols, sheet_name):
    industry = Industry(spreadsheet, industry, cols)
    industry.summary()
    industry.write_headers()
    industry.write_spreadsheet()
    industry.save_wb(sheet_name)


SUM_HEADERS = ['Previous Close', 'Open', 'Bid', 'Ask', "Day's Range", '52 Week Range', 'Volume', 'Avg. Volume',
               'Market Cap', 'Beta (5Y Monthly)', 'PE Ratio (TTM)', 'EPS (TTM)', 'Earnings Date',
               'Forward Dividend & Yield', 'Ex-Dividend Date', '1y target Est.']
STAT_HEADERS = []


class Industry:
    def __init__(self, filename, ind, cols):
        self.filename = filename
        self.industry = ind
        self.stocks = []
        self.headers = []

        self.new_wb = Workbook()
        self.new_ws = self.new_wb.create_sheet(self.industry)

        pd.options.display.float_format = '{:.0f}'.format
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb[self.industry]

        self.read_spreadsheet()
        self.read_columns(cols)

    def read_columns(self, cols):
        for row in self.ws.iter_rows(max_row=1, max_col=cols):  # TODO: optimize code
            for cell in row:
                self.headers.append(cell.value)

    def read_spreadsheet(self):
        for row in self.ws.iter_rows(min_row=2, max_col=3):
            stock = []
            for cell in row:
                stock.append(cell.value)
            self.stocks.append(stock)

        self.wb.close()

    def write_headers(self):
        for i in range(len(self.headers)):
            self.new_ws.cell(row=1, column=(i + 1)).value = self.headers[i]

    def write_spreadsheet(self):
        for i in range(len(self.stocks)):
            for j in range(len(self.stocks[i])):
                self.new_ws.cell(row=(i + 2), column=(j + 1)).value = self.stocks[i][j]

    def save_wb(self, file):
        self.new_wb.save(file)

    def summary(self):
        for stock in self.stocks:
            get_summary(stock)

        list.extend(self.headers, SUM_HEADERS)

    def update_stats(self):
        for stock in self.stocks:
            get_stats(stock)

        list.extend(self.headers, STAT_HEADERS)


stock = ['Canadian Utilities', 'CU.TO', 'Toronto Stock Exchange']
get_summary(stock)
print(stock)
