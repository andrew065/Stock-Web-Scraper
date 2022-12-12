import csv
import os
import time
import selenium.common.exceptions
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
from lxml import html


def get_info():
    stock_spreadsheet = load_workbook(filename='Spreadsheets/Wharton Stock List.xlsx')
    industries = stock_spreadsheet.sheetnames

    all_stocks = {}
    for industry in industries[1:]:
        stocks = []
        for row in stock_spreadsheet[industry].iter_rows(min_row=2, max_col=2):
            stock = []
            for cell in row:
                stock.append(cell.value)
            stocks.append(stock)
        all_stocks[industry] = stocks

    return all_stocks


driver = webdriver.Chrome()
industry_stocks = get_info()

for industry in industry_stocks:
    for i in range(len(industry)):
        try:
            data = industry_stocks[industry][i]
            url = "https://finance.yahoo.com/quote/%s/key-statistics?p=%s" % (data[1], data[1])
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'lxml')
            stats = str(soup.find_all("table", class_="W(100%) Bdcl(c)"))
            s = """Held by Institutions</span> <sup aria-label="Data provided by Refinitiv.">1</sup></td><td class="Fw(500) Ta(end) Pstart(10px) Miw(60px)">"""
            try:
                index = stats.index(s) + len(s)
                industry_stocks[industry][i].append(stats[index: index + 6])
            except ValueError:
                industry_stocks[industry][i].append(0)
        except IndexError:
            pass

print(industry_stocks)

for industry in industry_stocks:
    with open(f'{industry}.csv', 'w', newline='') as csvfile:
        fieldnames = ['Company', 'Ticker', '% Held by Institutions']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writerow({'Company': 'Company', 'Ticker': 'Ticker', '% Held by Institutions': '% Held by Institutions'})

        for key in industry_stocks[industry]:
            if len(key) == 3:
                writer.writerow({'Company': key[0], 'Ticker': key[1], '% Held by Institutions': key[2]})

        csvfile.close()


driver.quit()
