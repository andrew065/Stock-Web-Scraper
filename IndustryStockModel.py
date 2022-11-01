import selenium.common.exceptions
from selenium import webdriver
from bs4 import BeautifulSoup
import time, os, csv
from openpyxl import load_workbook


def download_stock_info(name, ticker):
    url = 'https://www.macrotrends.net/stocks/charts/%s/%s/stock-price-history'% (ticker, name)
    driver.get(url)
    try:
        driver.switch_to.frame('chart_iframe')
    except selenium.common.exceptions.NoSuchFrameException as error:
        print(error, name, ticker)
        return 0
    except selenium.common.exceptions.NoSuchElementException as error:
        print(error, name, ticker)
        return 0
    soup = BeautifulSoup(driver.page_source, 'lxml')

    scripts = soup.findAll("script")
    script = scripts[-1].string
    link = script[script.find('window.parent.location.href') + 26:
                  script.find('$(".imageExport").click( function()')].split('\'')[1]
    driver.get(link)
    time.sleep(0.1)


def format_stock_name(stock_name):
    words = stock_name.lower().split(' ')
    final_name = ''
    for word in words:
        if word not in ['co', 'inc', 'ltd', 'corp', '&', 'group', 'holdings', 'sa', 'plc']:
            if final_name == '':
                final_name += word
            else:
                final_name += '-' + word

    while '&' in final_name:
        final_name = final_name.replace('&', '-')
    return final_name


def get_company_info():
    stock_spreadsheet = load_workbook(filename='Spreadsheets/Wharton Stock List.xlsx')
    industries = stock_spreadsheet.sheetnames

    all_stocks = {}
    for industry in industries[1:]:
        stocks = []
        for row in stock_spreadsheet[industry].iter_rows(min_row=2, max_col=2):
            stock = []
            for cell in row:
                stock.append(cell.value)
            stocks.append(stock)
        all_stocks[industry] = stocks

    return all_stocks


def scrape_macrotrends_stockprice():
    for industry in industry_stocks:
        print('\n', industry)
        for stock in industry_stocks[industry]:
            if 'SA' in stock[0]:
                print(stock[0])
                result = download_stock_info(format_stock_name(stock[0]), stock[1])
                if result == 0:
                    print(format_stock_name(stock[0]), stock[1])


def read_industry_stock_data():
    for industry in industry_stocks:
        directory = f"Industry Stock Data/{industry}"

        for filename in os.listdir(directory):
            file = os.path.join(directory, filename)

            if os.path.isfile(file):
                with open(file, newline='') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        download_stock_info(format_stock_name(row['Name']), row['Symbol'])


driver = webdriver.Chrome()
industry_stocks = get_company_info()

read_industry_stock_data()

driver.close()
